import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D #3次元データの表示
import openpyxl
import numpy as np

'''表示用関数'''  
def scatter_data_3d(data,target,fn="fig",view=None): #描写 x,y,重心,ファイル名
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    ax.set_xlabel("feature1")
    ax.set_ylabel("feature2")
    ax.set_zlabel("output")
    ax.plot(data.T[0],data.T[1],target,marker="o",linestyle='None')
    if(view is not None):
      ax.view_init(elev=view[0], azim=view[1])
    plt.title(fn)
    plt.savefig(fn+".png")
    plt.close()

def scatter_2data_3d(data,y_true,y_pred,fn="fig",view=[10,45]): #view=[横方向,上方向]
  fig = plt.figure()
  ax = fig.add_subplot(111, projection='3d')
  ax.set_xlabel("feature1")
  ax.set_ylabel("feature2")
  ax.set_zlabel("output")
  ax.plot(data.T[0],data.T[1],y_true,marker="o",linestyle='None',label="y_true")
  ax.plot(data.T[0],data.T[1],y_pred,marker="x",linestyle='None',label="y_pred")
  if(view is not None):
    ax.view_init(elev=view[0], azim=view[1])
  plt.legend(loc="upper left")
  plt.title(fn)
  plt.savefig(fn+".png")
  plt.close()

def tar_to_color(target): #整数を色に変える変数
  if(type(target[0]) is np.ndarray):
    return target
  else:
    rtn = ["rbg"[x] for x in target]
    return rtn

def int_to_color(target): #整数を色に変える関数
  return "rbg"[target]

def scatter_data(data,target,fn="fig"): #描写 x,y,重心,ファイル名
  plt.figure()
  plt.scatter(data,target)
  plt.title(fn)
  plt.savefig("imgs/"+fn+".png")

def save_excel(data,columns=None,file_name ="a.xlsx",sheet_name=None): #excelファイルの書き込み
  path = "/content/drive/MyDrive/experiment/"+file_name
  wb = openpyxl.load_workbook(path)
  ws = wb.create_sheet()
  if(sheet_name is not None):
    ws.title = sheet_name
  if(columns is not None):
    data.insert(0,columns)
  if(len(data)==1):
    for i in range(len(data)):
      ws.cell(i+1,1,value=data[i])
  else:
    for d in range(len(data)):
      for dd in range(len(data[d])):
        ws.cell(d+1,dd+1,value = data[d][dd])
  wb.save(path)
